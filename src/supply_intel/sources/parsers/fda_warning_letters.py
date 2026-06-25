from __future__ import annotations

import csv
import hashlib
from datetime import UTC, datetime
from io import BytesIO, StringIO

from openpyxl import load_workbook  # type: ignore[import-untyped]

from supply_intel.models.documents import DocumentChunk
from supply_intel.models.source import RawDocument


def parse_fda_warning_letters_document(document: RawDocument) -> list[DocumentChunk]:
    rows = _rows_from_document(document)
    chunks: list[DocumentChunk] = []
    for index, row in enumerate(rows):
        if not row.get("company_name") and not row.get("subject"):
            continue
        text = _row_text(row)
        chunks.append(
            DocumentChunk(
                raw_document_id=document.id,
                chunk_index=index,
                chunk_type="tabular_record",
                title=row.get("company_name") or row.get("subject"),
                text=text,
                structured_data=row,
                content_hash=hashlib.sha256(text.encode("utf-8")).hexdigest(),
            )
        )
    return chunks


def _rows_from_document(document: RawDocument) -> list[dict[str, str]]:
    if document.payload_bytes and _looks_like_xlsx(document):
        return _rows_from_xlsx(document.payload_bytes)
    if document.payload_text:
        return _rows_from_csv(document.payload_text)
    return []


def _looks_like_xlsx(document: RawDocument) -> bool:
    payload = document.payload_bytes
    if payload is None:
        return False
    content_type = document.content_type or ""
    return (
        "spreadsheetml.sheet" in content_type
        or "excel" in content_type
        or payload.startswith(b"PK")
    )


def _rows_from_xlsx(payload: bytes) -> list[dict[str, str]]:
    workbook = load_workbook(BytesIO(payload), read_only=True, data_only=True)
    worksheet = workbook.active
    raw_headers: list[object] | None = None
    rows: list[dict[str, str]] = []
    for values in worksheet.iter_rows(values_only=True):
        if raw_headers is None:
            raw_headers = list(values)
            continue
        row = _row_from_values(raw_headers, list(values))
        if row:
            rows.append(row)
    workbook.close()
    return rows


def _rows_from_csv(payload: str) -> list[dict[str, str]]:
    reader = csv.DictReader(StringIO(payload))
    return [
        {
            _normalize_header(key): _cell_text(value)
            for key, value in row.items()
            if key is not None and _normalize_header(key)
        }
        for row in reader
    ]


def _row_from_values(headers: list[object], values: list[object]) -> dict[str, str]:
    row: dict[str, str] = {}
    for index, header in enumerate(headers):
        normalized = _normalize_header(_cell_text(header))
        if not normalized:
            continue
        value = values[index] if index < len(values) else None
        rendered = _cell_text(value)
        if rendered:
            row[normalized] = rendered
    return row


def _normalize_header(value: str) -> str:
    normalized = "_".join(part for part in value.casefold().replace("/", " ").split() if part)
    aliases = {
        "posted_date": "posted_date",
        "letter_issue_date": "letter_issue_date",
        "company_name": "company_name",
        "issuing_office": "issuing_office",
        "subject": "subject",
        "response_letter": "response_letter_url",
        "response_letter_url": "response_letter_url",
        "closeout_letter": "closeout_letter_url",
        "close_out_letter": "closeout_letter_url",
        "closeout_letter_url": "closeout_letter_url",
    }
    return aliases.get(normalized, normalized)


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        if value.tzinfo:
            return value.astimezone(UTC).date().isoformat()
        return value.date().isoformat()
    return str(value).strip()


def _row_text(row: dict[str, str]) -> str:
    fields = [
        f"Posted date: {row.get('posted_date', '')}",
        f"Letter issue date: {row.get('letter_issue_date', '')}",
        f"Company name: {row.get('company_name', '')}",
        f"Issuing office: {row.get('issuing_office', '')}",
        f"Subject: {row.get('subject', '')}",
        f"Response letter: {row.get('response_letter_url', '')}",
        f"Closeout letter: {row.get('closeout_letter_url', '')}",
    ]
    return "\n".join(field for field in fields if not field.endswith(": "))
