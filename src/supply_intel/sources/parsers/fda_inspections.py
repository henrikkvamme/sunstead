from __future__ import annotations

import csv
import hashlib
import re
from datetime import UTC, datetime
from io import BytesIO, StringIO

from openpyxl import load_workbook  # type: ignore[import-untyped]

from supply_intel.models.documents import DocumentChunk
from supply_intel.models.source import RawDocument


def parse_fda_inspections_dashboard_document(document: RawDocument) -> list[DocumentChunk]:
    rows = _rows_from_document(document)
    chunks: list[DocumentChunk] = []
    for index, row in enumerate(rows):
        if not _has_inspection_signal(row):
            continue
        row["dataset_type"] = _dataset_type(row)
        text = _row_text(row)
        chunks.append(
            DocumentChunk(
                raw_document_id=document.id,
                chunk_index=index,
                chunk_type="tabular_record",
                title=_row_title(row),
                text=text,
                structured_data=row,
                content_hash=hashlib.sha256(text.encode("utf-8")).hexdigest(),
            )
        )
    return chunks


def _rows_from_document(document: RawDocument) -> list[dict[str, str]]:
    if document.payload_bytes and _looks_like_xlsx(document):
        return _rows_from_xlsx(document.payload_bytes)
    if document.payload_text and not _looks_like_dashboard_html(document.payload_text):
        return _rows_from_csv(document.payload_text)
    return []


def _looks_like_xlsx(document: RawDocument) -> bool:
    payload = document.payload_bytes
    if payload is None:
        return False
    content_type = document.content_type or ""
    return (
        "spreadsheetml.sheet" in content_type
        or "excel" in content_type
        or payload.startswith(b"PK")
    )


def _looks_like_dashboard_html(payload: str) -> bool:
    sample = payload[:512].casefold()
    return "<!doctype html" in sample or "<html" in sample


def _rows_from_xlsx(payload: bytes) -> list[dict[str, str]]:
    workbook = load_workbook(BytesIO(payload), read_only=True, data_only=True)
    rows: list[dict[str, str]] = []
    try:
        for worksheet in workbook.worksheets:
            raw_headers: list[object] | None = None
            for values in worksheet.iter_rows(values_only=True):
                if raw_headers is None:
                    raw_headers = list(values)
                    continue
                row = _row_from_values(raw_headers, list(values))
                if row:
                    row.setdefault("source_sheet", worksheet.title)
                    rows.append(row)
    finally:
        workbook.close()
    return rows


def _rows_from_csv(payload: str) -> list[dict[str, str]]:
    reader = csv.DictReader(StringIO(payload))
    return [
        {
            _normalize_header(key): _cell_text(value)
            for key, value in row.items()
            if key is not None and _normalize_header(key)
        }
        for row in reader
    ]


def _row_from_values(headers: list[object], values: list[object]) -> dict[str, str]:
    row: dict[str, str] = {}
    for index, header in enumerate(headers):
        normalized = _normalize_header(_cell_text(header))
        if not normalized:
            continue
        value = values[index] if index < len(values) else None
        rendered = _cell_text(value)
        if rendered:
            row[normalized] = rendered
    return row


def _normalize_header(value: str) -> str:
    normalized = re.sub(r"[^a-z0-9]+", "_", value.casefold()).strip("_")
    aliases = {
        "action_type": "classification",
        "city": "city_name",
        "company_name": "legal_name",
        "country": "country_name",
        "document_url": "form_483_url",
        "fda_483_url": "form_483_url",
        "fd_c_act_section": "fdca_section",
        "fdca": "fdca_section",
        "fdca_reference": "fdca_section",
        "fei": "fei_number",
        "fei_no": "fei_number",
        "firm_name": "legal_name",
        "form_483_url": "form_483_url",
        "inspection_classification": "classification",
        "inspectionid": "inspection_id",
        "legalname": "legal_name",
        "long_desc": "long_description",
        "publish_date": "publish_date",
        "published_date": "publish_date",
        "short_desc": "short_description",
        "state": "inspection_state",
        "state_code": "inspection_state",
        "url": "form_483_url",
    }
    return aliases.get(normalized, normalized)


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        if value.tzinfo:
            return value.astimezone(UTC).date().isoformat()
        return value.date().isoformat()
    return str(value).strip()


def _has_inspection_signal(row: dict[str, str]) -> bool:
    return any(
        row.get(field)
        for field in (
            "inspection_id",
            "citation_id",
            "fei_number",
            "legal_name",
            "classification",
            "short_description",
            "record_type",
            "publish_date",
        )
    )


def _dataset_type(row: dict[str, str]) -> str:
    explicit = row.get("record_type", "").casefold()
    sheet = row.get("source_sheet", "").casefold()
    if "citation" in explicit or "citation" in sheet or row.get("citation_id"):
        return "citation"
    if "483" in explicit or "483" in sheet or row.get("publish_date") or row.get("form_483_url"):
        return "published_483"
    return "inspection"


def _row_title(row: dict[str, str]) -> str:
    dataset_type = row.get("dataset_type", "inspection")
    legal_name = row.get("legal_name") or row.get("fei_number") or "Unknown firm"
    inspection_id = row.get("inspection_id")
    if dataset_type == "citation":
        return f"FDA inspection citation: {row.get('citation_id') or inspection_id or legal_name}"
    if dataset_type == "published_483":
        return f"FDA published 483: {inspection_id or legal_name}"
    return f"FDA inspection: {inspection_id or legal_name}"


def _row_text(row: dict[str, str]) -> str:
    fields = [
        f"Dataset type: {row.get('dataset_type', '')}",
        f"Inspection ID: {row.get('inspection_id', '')}",
        f"Citation ID: {row.get('citation_id', '')}",
        f"FEI number: {row.get('fei_number', '')}",
        f"Legal name: {row.get('legal_name', '')}",
        f"City name: {row.get('city_name', '')}",
        f"Inspection state: {row.get('inspection_state', '')}",
        f"Country name: {row.get('country_name', '')}",
        f"Fiscal year: {row.get('fiscal_year', '')}",
        f"Product type: {row.get('product_type', '')}",
        f"Program area: {row.get('program_area', '')}",
        f"Project area: {row.get('project_area', '')}",
        f"Classification: {row.get('classification', '')}",
        f"Inspection end date: {row.get('inspection_end_date', '')}",
        f"Posted citations: {row.get('posted_citations', '')}",
        f"Short description: {row.get('short_description', '')}",
        f"Long description: {row.get('long_description', '')}",
        f"CFR section: {row.get('cfr_section', '')}",
        f"FDCA section: {row.get('fdca_section', '')}",
        f"Record date: {row.get('record_date', '')}",
        f"Publish date: {row.get('publish_date', '')}",
        f"Form 483 URL: {row.get('form_483_url', '')}",
        f"Additional details: {row.get('additional_details', '')}",
    ]
    return "\n".join(field for field in fields if not field.endswith(": "))
