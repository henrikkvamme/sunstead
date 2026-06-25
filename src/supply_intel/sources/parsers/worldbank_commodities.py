from __future__ import annotations

import csv
import hashlib
import io
from typing import Any

from openpyxl import load_workbook  # type: ignore[import-untyped]

from supply_intel.models.documents import DocumentChunk
from supply_intel.models.source import RawDocument

MONTHLY_PRICES_SHEET = "Monthly Prices"
COMMODITY_ROW = 5
UNIT_ROW = 6
FIRST_DATA_ROW = 7


def parse_worldbank_commodity_prices_document(document: RawDocument) -> list[DocumentChunk]:
    records = _price_records(document)
    chunks: list[DocumentChunk] = []
    for index, record in enumerate(records):
        commodity = str(record.get("commodity_name", "")).strip()
        period = str(record.get("period", "")).strip()
        if not commodity or not period:
            continue
        text = _price_text(record)
        chunks.append(
            DocumentChunk(
                raw_document_id=document.id,
                chunk_index=index,
                chunk_type="commodity_price_observation",
                title=f"{commodity} {period}",
                text=text,
                structured_data=record,
                content_hash=hashlib.sha256(text.encode("utf-8")).hexdigest(),
            )
        )
    return chunks


def _price_records(document: RawDocument) -> list[dict[str, object]]:
    if document.payload_bytes:
        return _records_from_xlsx(document.payload_bytes)
    if document.payload_text:
        return _records_from_csv(document.payload_text)
    return []


def _records_from_xlsx(payload: bytes) -> list[dict[str, object]]:
    workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    if MONTHLY_PRICES_SHEET not in workbook.sheetnames:
        return []
    sheet = workbook[MONTHLY_PRICES_SHEET]
    updated_on = str(sheet.cell(row=4, column=1).value or "").strip()
    commodity_names = [
        str(value or "").strip()
        for value in next(
            sheet.iter_rows(
                min_row=COMMODITY_ROW,
                max_row=COMMODITY_ROW,
                values_only=True,
            )
        )
    ]
    units = [
        str(value or "").strip()
        for value in next(sheet.iter_rows(min_row=UNIT_ROW, max_row=UNIT_ROW, values_only=True))
    ]
    latest_row = _latest_numeric_row(sheet)
    if latest_row is None:
        return []
    period = str(latest_row[0] or "").strip()
    records: list[dict[str, object]] = []
    for index, raw_value in enumerate(latest_row[1:], start=1):
        value = _float_value(raw_value)
        commodity_name = commodity_names[index] if index < len(commodity_names) else ""
        if value is None or not commodity_name:
            continue
        raw_unit = units[index] if index < len(units) else ""
        records.append(
            {
                "period": period,
                "commodity_name": commodity_name,
                "raw_unit": raw_unit,
                "value": value,
                "updated_on": updated_on,
                "source_table": MONTHLY_PRICES_SHEET,
            }
        )
    return records


def _latest_numeric_row(sheet: Any) -> tuple[object, ...] | None:
    latest: tuple[object, ...] | None = None
    for row in sheet.iter_rows(min_row=FIRST_DATA_ROW, values_only=True):
        period = str(row[0] or "").strip()
        if not period:
            continue
        if any(_float_value(value) is not None for value in row[1:]):
            latest = tuple(row)
    return latest


def _records_from_csv(payload: str) -> list[dict[str, object]]:
    rows = csv.DictReader(io.StringIO(payload))
    records: list[dict[str, object]] = []
    for row in rows:
        value = _float_value(row.get("value"))
        commodity_name = str(row.get("commodity_name") or "").strip()
        period = str(row.get("period") or "").strip()
        if value is None or not commodity_name or not period:
            continue
        records.append(
            {
                "period": period,
                "commodity_name": commodity_name,
                "raw_unit": str(row.get("unit") or row.get("raw_unit") or "").strip(),
                "value": value,
                "updated_on": str(row.get("updated_on") or "").strip(),
                "source_table": str(row.get("source_table") or MONTHLY_PRICES_SHEET).strip(),
            }
        )
    return records


def _float_value(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, int | float):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text or text in {"…", "..", "NA", "N/A", "-"}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _price_text(record: dict[str, Any]) -> str:
    fields = [
        f"Commodity: {record.get('commodity_name', '')}",
        f"Period: {record.get('period', '')}",
        f"Value: {record.get('value', '')}",
        f"Unit: {record.get('raw_unit', '')}",
        f"Updated on: {record.get('updated_on', '')}",
        f"Source table: {record.get('source_table', '')}",
    ]
    return "\n".join(field for field in fields if not field.endswith(": "))
